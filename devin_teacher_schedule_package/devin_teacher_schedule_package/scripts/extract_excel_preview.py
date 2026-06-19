from pathlib import Path
import json
import csv

# Optional helper script for Devin / Windsurf.
# Run from the project root:
#   python scripts/extract_excel_preview.py

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("Please install openpyxl first: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "data" / "2026_TEACHER_SCHEDULES.xlsx"
OUTPUT = ROOT / "output"
OUTPUT.mkdir(exist_ok=True)

wb = load_workbook(INPUT, data_only=True)
summary = []

for ws in wb.worksheets:
    sheet_name_safe = "".join(ch if ch.isalnum() else "_" for ch in ws.title).strip("_") or "sheet"
    rows = []
    for row in ws.iter_rows():
        values = [cell.value for cell in row]
        # Keep rows that have at least one visible value
        if any(v is not None and str(v).strip() != "" for v in values):
            rows.append(values)

    csv_path = OUTPUT / f"{sheet_name_safe}.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    json_path = OUTPUT / f"{sheet_name_safe}.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2, default=str)

    summary.append({
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "non_empty_rows_exported": len(rows),
        "csv": str(csv_path.relative_to(ROOT)),
        "json": str(json_path.relative_to(ROOT)),
    })

with (OUTPUT / "workbook_summary.json").open("w", encoding="utf-8") as f:
    json.dump(summary, f, ensure_ascii=False, indent=2)

print(json.dumps(summary, ensure_ascii=False, indent=2))
